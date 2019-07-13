'''
Created on May 14th, 2018
Author: Miles McCall
Sources:
Description: Parse the "Pollinator Plant Master List" maintained by Signe and
            generate a formatted version for uploading to the website.
'''

# External imports
import os
import sys
import string
import errno
import csv
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
# Local imports
import col_functions
import file_functions
from file_functions import test_import

#Functions

def letter_to_index(letter):
    """Converts a column letter, e.g. "A", "B", "AA", "BC" etc. to a zero based
    column index.

    A becomes 0, B becomes 1, Z becomes 25, AA becomes 26 etc.

    Args:
        letter (str): The column index letter.
    Returns:
        The column index as an integer.
    """
    letter = letter.upper()
    result = 0

    for index, char in enumerate(reversed(letter)):
        # Get the ASCII number of the letter and subtract 64 so that A
        # corresponds to 1.
        num = ord(char) - 64

        # Multiply the number with 26 to the power of `index` to get the correct
        # value of the letter based on it's index in the string.
        final_num = (26 ** index) * num

        result += final_num

    # Subtract 1 from the result to make it zero-based before returning.
    return result - 1

def count_rows(workbook, worksheet):
    wb = load_workbook(str(workbook))
    ws = wb[str(worksheet)]
    row_count = str(ws.max_row)
    print("count rows:",row_count)
    return row_count

def count_cols(workbook, worksheet):
    wb = load_workbook(str(workbook))
    ws = wb[str(worksheet)]
    col_count = str(ws.max_column)
    print("count cols:",col_count)
    return col_count


def parse_cmd_line():
    # Parse the command line arguments:
    # Determine an input path, output path, and input file type
    print("Parsing command line arguments...")

    # Init vars
    i = 0
    in_file = ''
    out_file = ''
    in_type = ''

    # Iterate through cmd line and assign strings to input and output paths
    for arg in sys.argv:
        if arg == "--input":
            in_file = sys.argv[i+1]
        elif arg == "--output":
            out_file = sys.argv[i+1]
        i += 1

    # Input path:
        # data/folder_name/file_name

    # Input file is required
    if in_file == '':
        print("\'--input\' argument must be provided. Exitting.")
        sys.exit()

    # The input file must be kept in data dir
    if in_file.split('/')[0] != "data":
        print("\'--input\' file must be saved inside the data directory. Exitting.")
        sys.exit()

    # Check input file type
    in_type = in_file.split('/')[len(in_file.split('/')) - 1]
    in_type = in_type.split(".")[len(in_type.split('.')) - 1]

    # Output path:
        # results/folder_name/file_name

    # If the output file does exist it must be kept in the results dir
    if out_file != '' and out_file.split('/')[0] != "results":
        print("\'--output\' file must be saved inside the results directory. Exitting.")
        sys.exit()

    # If output was not specified, use the input folder name
    if out_file == '':
        # We will use the split file path components
        out_file = "results/" + in_file.split('/')[1] + "/results.csv"
        out_file_windows = "results/" + in_file.split('/')[1] + "/results_windows.csv"

    # Its necessary to append a '/' to the output folder so its treated as a dir
    out_folder = "results/" + in_file.split('/')[1] + "/"

    # Create output folder
    if not os.path.exists(os.path.dirname(out_folder)):
        try:
            os.makedirs(os.path.dirname(out_folder))
        except OSError as exc: # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise

    # Create output file
    f = open(out_file, "w")

    # Return vars
    return in_file, in_type.lower(), out_file, out_file_windows

def read_csv_header(file_string):
    file = open(file_string, "r") # Open CSV file
    return file.readline()

def read_xlsx_header(wb_name, ws_name):
    wb = load_workbook(wb_name)
    ws = wb[str(ws_name)]

    max_col = list(string.ascii_lowercase)[ws.max_column - 1].upper()
    print(max_col)
    print(ws.max_column)

    selection_str = "A1:" + max_col + "1"
    return ws[selection_str].ecnode("utf8")

def read_csv(file_string):
    print("\tReading from CSV...")

    # Open CSV file
    file = open(file_string, "r")

    # Read the first line into header var
    header_row = file.readline()

    # Iterate through rest of file, saving in array
    file_rows = []
    for line in file:
        file_rows.append(line)

    # Returns the header row and line array
    return header_row, file_rows

def read_xlsx(file_string):
    print("\tReading from Excel spreadsheet...")

def read_data(file_string, file_type):
    print("Reading data from input source...")
    # Variables to capture the header row and following data
    header = ''
    file_rows = []

    # Check which file type to read from
    if file_type == "csv":
        header, file_rows = read_csv(file_string)
    elif file_type == "xlsx":
        header, file_rows = read_xlsx(file_string)
    else:
        print("Invalid input file type. Exitting.")
        sys.exit()

    # Strip header of extra characters and converts individuals chars into words
    header = header.strip().split(',')
    return header, file_rows

def search_header(header, search_str):
    # Locate the index of a string in the header row
    for index, col in enumerate(header):
        if col == search_str:
            return index

def print_out_header(line_to_print, csv_file):
    print(line_to_print)
    with open(csv_file, 'w', newline = '') as file:
        for index, col in enumerate(line_to_print):
            # If the current col has a value, print it
            if col != '':
                file.write(str(col))
            # If the current col isn't the last, print a comma
            # This can be separated because both
                #blank and full cols need a comma
            if index < len(line_to_print):
                file.write(',')
        # Append newline as the last step to start the next row
        file.write("\n")

def print_out_row(line_to_print, csv_file):
    #print(line_to_print)
    with open(csv_file, 'a', newline = '') as f:
        writer = csv.writer(f)
        writer.writerow(line_to_print)

    '''
    with open(csv_file,'a',newline='') as file:
        # Append generated row to output file
        print(repr(line_to_print))
        for index, col in enumerate(line_to_print):
            # If the current col has a value, print it
            if col != '':
                file.write(str(col))
            # If the current col isn't the last, print a comma
            # This can be separated because both
                #blank and full cols need a comma
            if index < len(line_to_print):
                file.write(',')
        # Append newline as the last step to start the next row
        #file.write("\n")
    '''

def remove_blank_rows(out_file):
    header, rows = read_csv(out_file)
    print_out_header(header,out_file)
    for line in rows:
        a = repr(str(line))
        #print("line:",a,"\n\t")
        if line == "\n":
            print("blank row removed")
        else:
            #print("keeping line")
            print_out_row(line,out_file)

def check_for_cols(in_header, in_row, query_string):
    search_res = search_header(in_header,query_string)
    #print(search_res)
    if search_res is None:
        return ''
    else:
        return in_row[search_res]

def gen_output(out_header, out_file, in_header, in_data):
    # Create rows of formatted data and append to output csv
    print("Generating output data...")

    # Print header row
    print_out_header(out_header,out_file)

    # Parse input rows
    i = 0
    for in_row in csv.reader(in_data, skipinitialspace=True):
        i += 1
        print("\nrow:\n",in_row)

        # Init the output row
        out_row = []

        # Date Label Printed
        # Date Label Sent
        # Observation No.
        # Voucher No.
        out_row.append(" ")
        out_row.append(" ")
        out_row.append(" ")
        out_row.append(" ")

        # iNaturalist ID
        id = check_for_cols(in_header, in_row, "user_id")
        out_row.append(id)

        # iNaturalist Alias
        iNat_alias = check_for_cols(in_header, in_row, "user_login")
        out_row.append(iNat_alias)

        # Collector - First Name
        # Collector - First Initial
        # Collector - Last Name
        u_name = check_for_cols(in_header, in_row, "user_login")
        f_name, f_initial, l_name = col_functions.collector_name("data/usernames.csv",u_name)
        out_row.append(f_name)
        out_row.append(f_initial)
        out_row.append(l_name)

        # Sample ID
        sampleid = check_for_cols(in_header, in_row, "field:sample id.")
        out_row.append(sampleid)

        # Specimen ID
        specimenid = check_for_cols(in_header, in_row, "field:number of bees collected")
        out_row.append(specimenid)


        # Collection Day 1
        # Month 1
        # Year 1
        # Time 1
        date1 = check_for_cols(in_header, in_row, "observed_on")
        day1, month1, year1 = col_functions.date_1(date1)
        time1 = col_functions.time_1(in_row[search_header(in_header,"time_observed_at")])
        out_row.append(day1)
        out_row.append(month1)
        out_row.append(year1)
        out_row.append(time1)

        # Collection Day 2
        # Moth 2
        # Year 2
        # Day 2 merge
        # Time 2
        date2 = check_for_cols(in_header, in_row, "field:trap removed")
        day2, month2, year2, merge2 = col_functions.date_2(date2)
        time2 = col_functions.time_2(date2)
        out_row.append(day2)
        out_row.append(month2)
        out_row.append(year2)
        out_row.append(merge2)
        out_row.append(time2)

        # Country
        country = "USA"
        out_row.append(country)

        # State
        state = check_for_cols(in_header, in_row, "place_state_name")
        if state == "Oregon":
            state = "OR"
        out_row.append(state)

        # County
        county = check_for_cols(in_header, in_row, "place_county_name")
        out_row.append(county)

        # Location
        # Abbreviated Location
        place_guess = check_for_cols(in_header, in_row, "place_guess")
        location = col_functions.location_guess(place_guess,"data/OR_cities.csv")
        abbreviated_location = ''
        out_row.append(location)
        out_row.append(abbreviated_location)

        # Dec. Lat.
        # Dec. Long.
        lat = check_for_cols(in_header, in_row, "latitude")
        long = check_for_cols(in_header, in_row, "longitude")
        if lat == '' or long == '':
            out_row.append('')
            out_row.append('')
        else:
            lat = col_functions.round_coord(lat)
            long = col_functions.round_coord(long)
            if lat is None or long is None:
                out_row.append('')
                out_row.append('')
            else:
                out_row.append(lat)
                out_row.append(long)

        # Pos Accuracy
        pos_acc = check_for_cols(in_header, in_row, "positional_accuracy")
        out_row.append(pos_acc)

        # Elevation
        if lat is None or long is None or lat == '' or long == '':
            out_row.append('')
        else:
            elevation = col_functions.elevation(lat,long)
            out_row.append(elevation)

        # Collection method
        collection_method = check_for_cols(in_header, in_row, "field:oba collection method")
        out_row.append(collection_method)

        # Associated plant - family
        # Associated plant - species
        # Associated plant - iNaturalist url
        family = check_for_cols(in_header, in_row, "taxon_family_name")
        species = check_for_cols(in_header, in_row, "scientific_name")
        url = check_for_cols(in_header, in_row, "url")
        out_row.append(family)
        out_row.append(species)
        out_row.append(url)
        # End of appending to output row

        # Append generated row to output file
        # If the row has multiple bees collected, expand by that many
        print("res:")
        if specimenid is not None:
            try:
                specimenid = int(specimenid)
                if int(specimenid) > 1:
                    print("multiple bees, printing",specimenid,"times...")
                    for i in range(1, int(specimenid)+1):
                        out_row[search_header(out_header,"Specimen ID")] = i
                        print_out_row(out_row,out_file)
                        print(out_row)
                else:
                    print_out_row(out_row,out_file)
                    print(out_row)
            except ValueError:
                pass  # it was a string, not an int.
                print_out_row(out_row,out_file)
                print(out_row)
        else:
            print_out_row(out_row,out_file)
            print(out_row)
        print()

def create_csv_windows(out_file, out_file_windows):
    #with open('/pythonwork/thefile_subset11.csv', 'w', newline='') as outfile:
    #writer = csv.writer(outfile)
    header, rows = read_csv(out_file)
    print_out_header(header,out_file_windows)
    for line in rows:
        temp = repr(str(line))
        #print(temp)
        temp = temp[:-3]
        #print(temp)
        print_out_row(temp,out_file_windows)

def main():
    # Intro
    print("iNaturalist Pipeline -----------------------")

    # Variables to keep track of
    input_file = ""
    output_file = ""
    input_file_type = ""

    # Parse command line arguments
    input_file, input_file_type, output_file, output_file_windows = parse_cmd_line()

    # Pipeline Description
    print("\tInput path:\t",input_file)
    print("\tInput type:\t",input_file_type)
    print("\tOutput path:\t",output_file)
    print("\tOutput path:\t",output_file_windows)
    print()

    # Choose which file reading function to call
    input_header, input_rows = read_data(input_file, input_file_type)
    print()

    # Sort columns before writing output
    output_header = "Date Label Printed,Date Label Sent,Observation No.,Voucher No.,iNaturalist ID,iNaturalist Alias,Collector - First Name,Collector - First Name Initial,Collector - Last Name,Sample ID,Specimen ID,Collection Day 1,Month 1,Year 1,Time 1,Collection Day 2,Month 2,Year 2,Collection Day 2 Merge,Time 2,Country,State,County,Location,Abbreviated Location,Dec. Lat.,Dec. Long.,Lat/Long Accuracy,Elevation,Collection method,Associated plant - family,Associated plant - species,Associated plant - Inaturalist URL".split(",")
    # Revisit
    #output_header2 = read_xlsx_header("data/4_16_19/Output_from_Script.xlsx","Sheet1")
    #print(output_header2)

    # Create output data
#gen_output(output_header, output_file, input_header, input_rows)
    #create_csv_windows(output_file, output_file_windows)
    print()

if __name__ == '__main__':
    main()
