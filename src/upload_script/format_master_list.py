'''
Created on May 14th, 2018
Author: Miles McCall
Sources:
Description: Parse the "Pollinator Plant Master List" maintained by Signe and
            generate a formatted version for uploading to the website.
'''

# External Imports
import os
import sys
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
# Local Imports

#Functions
def read_xlsx(wb_name, ws_name, min_col, min_row, max_col, max_row):
    # Initialize openpyxl vars
    wb = load_workbook(filename = str(wb_name))
    ws = wb[str(ws_name)]
    return_array = []

    # Iterate through xlsx file
    for row in ws.iter_rows(min_row, min_col, max_col,):
        temp_array = []
        for cell in row:
            temp_array.append(cell.value)
        return_array.append(temp_array)

    # Return array selection
    return return_array

def translate_format(master_array, input_wb, input_ws):
    # Load XLSX file
    wb = load_workbook(filename = str(input_wb))
    ws = wb[str(input_ws)]

    # translate master list rows to template format
    columns = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P',
            'Q','R','S','T','U','V','W','X','Y','Z']

    for row in master_array:
        result_array = []
        print(row)
        for col in columns:
            if(col == 'A'): # Title
                result_array.append('')
            elif(col == 'B'): # Icon
                result_array.append('')
            elif(col == 'C'): # Show Icon
                result_array.append('')
            elif(col == 'D'): # Featured Image
                result_array.append('')
            elif(col == 'E'): # Alternative text
                result_array.append('')
            elif(col == 'F'): # Featured Image Title
                result_array.append('')
            elif(col == 'G'): # Image attribute
                result_array.append('')
            elif(col == 'H'): # Common Name
                result_array.append(str(row[2]))
            elif(col == 'I'): # Genus and species
                temp = str(row[0]) + " " + str(row[1])
                result_array.append(temp)
            elif(col == 'J'): # Family
                result_array.append(str(row[4]))
            elif(col == 'K'): # Habitat and Cultivation header
                result_array.append('Habitat value and plant care')
            elif(col == 'L'): # Habitat Teaser
                result_array.append('Pollinators and wildife attracted, bloom details, sun & soil')
            elif(col == 'M'): # Growth habit
                if('ann' in str(row[5]).lower()):
                    result_array.append('Annual')
                elif('per' in str(row[5]).lower()):
                    result_array.append('Perennial')
                elif('sh' in str(row[5]).lower()):
                    result_array.append('Shrub')
                elif('tr' in str(row[5]).lower()):
                    result_array.append('Tree')
                elif('v' in str(row[5]).lower()):
                    result_array.append('Vine')
            elif(col == 'N'): # Oregon native plant
                if 'pnw' in str(row[25]).lower():
                    result_array.append('yes')
                elif('or' in str(row[25]).lower()):
                    result_array.append('yes')
                else:
                    result_array.append('no')
            elif(col == 'O'): # Edible
                result_array.append('manual entry')
            elif(col == 'P'): # Pollinators and Wildlife
                temp = ''; count = 0
                if 'x' in str(row[26]).lower():
                    temp += 'Honeybee'
                    count += 1
                if 'x' in str(row[27]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Bumblebee'
                    count += 1
                if 'x' in str(row[28]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Other native bee'
                    count += 1
                if 'hum' in str(row[30]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Hummingbird'
                    count += 1
                if 'mo' in str(row[30]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Moth'
                    count += 1
                if 'bu' in str(row[30]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Butterfly'
                    count += 1
                if '' not in str(row[32]):
                    if count > 0:
                        temp += ", "
                    temp += 'Larval Host'
                    count += 1
                result_array.append(temp)
            elif(col == 'Q'): # Susceptible to pests
                result_array.append('manual entry')
            elif(col == 'R'): # Flower color
                temp = ''; count = 0
                if 'bu' in str(row[18]).lower():
                    temp += 'Blue'
                    count += 1
                if 'gr' in str(row[18]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Green'
                    count += 1
                if 'lv' in str(row[18]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Lavender'
                    count += 1
                if 'or' in str(row[18]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Orange'
                    count += 1
                if 'pk' in str(row[18]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Pink'
                    count += 1
                if 'pp' in str(row[18]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Purple'
                    count += 1
                if 'rd' in str(row[18]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Red'
                    count += 1
                if 'wh' in str(row[18]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'White'
                    count += 1
                if 'yl' in str(row[18]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Yellow'
                    count += 1
                result_array.append(temp)
            elif(col == 'S'): # Bloom Season
                temp = ''; count = 0
                if 'early win' in str(row[17]).lower():
                    temp += 'Early Winter'
                    count += 1
                if 'mid win' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Mid Winter'
                    count += 1
                if 'late win' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Late Winter'
                    count += 1
                if 'early spr' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Early Spring'
                    count += 1
                if 'mid spr' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Mid Spring'
                    count += 1
                if 'late spr' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Late Spring'
                    count += 1
                if 'early sum' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Early Summer'
                    count += 1
                if 'mid sum' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Mid Summer'
                    count += 1
                if 'late sum' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Late Summer'
                    count += 1
                if 'early fall' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Early Fall'
                    count += 1
                if 'mid fall' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Mid Fall'
                    count += 1
                if 'late fall' in str(row[17]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Late Fall'
                    count += 1
                result_array.append(temp)
            elif(col == 'T'): # Attracts pollinators
                temp = 'no'
                if( '' not in str(row[26]).lower()       or '' not in str(row[27]).lower()
                    or '' not in str(row[28]).lower()    or '' not in str(row[29]).lower()
                    or '' not in str(row[30]).lower()
                ):
                    temp = 'yes'
                result_array.append(temp)
            elif(col == 'U'): # Problem plant for pol.
                result_array.append('manual entry')
            elif(col == 'V'): # Food resources
                temp = ''; count = 0
                if 'n' in str(row[24]).lower():
                    temp += 'Nectar'
                    count += 1
                if 'p' in str(row[24]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Pollen'
                    count += 1
                if 'f' in str(row[24]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Fruit'
                    count += 1
                if 'b' in str(row[24]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Berry'
                    count += 1
                result_array.append(temp)
            elif(col == 'W'): # Larval host for butterflies
                result_array.append('no')
            elif(col == 'X'): # Cultivation tolerances
                temp = ''; count = 0
                if 'sun' in str(row[22]).lower():
                    temp += 'Full sun'
                    count += 1
                if 'pt sun' in str(row[22]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Partial sun'
                    count += 1
                if 'dry' in str(row[23]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Dry soil'
                    count += 1
                if 'med' in str(row[23]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Medium moist soil'
                    count += 1
                if( 'moist' in str(row[23]).lower() or 'wet' in str(row[23]).lower() ):
                    if count > 0:
                        temp += ", "
                    temp += 'Moist soil'
                    count += 1
                result_array.append(temp)
            elif(col == 'Y'): # Availability
                temp = ''; count = 0
                if 'c' in str(row[36]).lower():
                    temp += 'Common'
                    count += 1
                if 'a' in str(row[36]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Available'
                    count += 1
                if 'u' in str(row[36]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Uncommon'
                    count += 1
                if 'h' in str(row[36]).lower():
                    if count > 0:
                        temp += ", "
                    temp += 'Hard'
                    count += 1
                result_array.append(temp)
            elif(col == 'Z'): # Grow Info
                result_array.append('manual entry')
        ws.append(result_array)
    wb.save(input_wb)

def main():
    # Init master vars
    master_wb = 'data/master_list.xlsx'
    master_ws = 'Master table'
    min_col = 'A';  min_row = '3'
    max_col = 'AK'; max_row = '282'
    # Extract values from master table
    master_res = read_xlsx(master_wb, master_ws, min_col, min_row, max_col, max_row)
    print(master_res)

    # Init template vars
    template_wb = 'data/plants_sample_data.xlsx'
    template_ws = 'Sample spreadsheet for the plan'

    translate_format(master_res, template_wb, template_ws)

if __name__ == '__main__':
    main()
