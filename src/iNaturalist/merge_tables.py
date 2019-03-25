'''
Created on May 14th, 2018
Author: Miles McCall
Sources:
Description: Parse the "Pollinator Plant Master List" maintained by Signe and
            generate a formatted version for uploading to the website.
'''

# External Imports
import os
import sys
import string
import errno
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
    #from openpyxl import Workbook
    #from openpyxl.reader.excel import load_workbook
    #from openpyxl.compat import range
    #from openpyxl.utils import get_column_letter

#Functions
def read_xlsx(wb_name, ws_name, min_col, min_row, max_col, max_row):
    # Initialize openpyxl vars
    wb = load_workbook(str(wb_name))
    ws = wb[str(ws_name)]
    return_array = []

    # Iterate through xlsx file
    selection_str = min_col + min_row + ':' + max_col + max_row
    for row in ws[selection_str]:
        temp_array = []
        for cell in row:
            temp_array.append(str(cell.value).encode("utf8"))
        if temp_array != [None]:
            return_array.append(temp_array)

    # Return array selection
    print(return_array)
    return return_array

def letter_to_index(letter):
    """Converts a column letter, e.g. "A", "B", "AA", "BC" etc. to a zero based
    column index.

    A becomes 0, B becomes 1, Z becomes 25, AA becomes 26 etc.

    Args:
        letter (str): The column index letter.
    Returns:
        The column index as an integer.
    """
    letter = letter.upper()
    result = 0

    for index, char in enumerate(reversed(letter)):
        # Get the ASCII number of the letter and subtract 64 so that A
        # corresponds to 1.
        num = ord(char) - 64

        # Multiply the number with 26 to the power of `index` to get the correct
        # value of the letter based on it's index in the string.
        final_num = (26 ** index) * num

        result += final_num

    # Subtract 1 from the result to make it zero-based before returning.
    return result - 1

def count_rows(workbook, worksheet):
    wb = load_workbook(str(workbook))
    ws = wb[str(worksheet)]
    row_count = str(ws.max_row)
    print("count rows:",row_count)
    return row_count

def count_cols(workbook, worksheet):
    wb = load_workbook(str(workbook))
    ws = wb[str(worksheet)]
    col_count = str(ws.max_column)
    print("count cols:",col_count)
    return col_count

# 'Eval' functions are called from the merge_tables function to evaluate the
# contents of each column for the output spreadsheet. For each row in the
# original table, each eval function is called sequentially to construct the
# desired output row.

# I chose to encapsulate each column value in its own eval function to make
# the interpretation of input data as modular as possible. Each eval function
# uses potentially different logic to generate properly formatted output values,
# and each function can be modified independently.

def eval_iNaturalistID(result_array, id_string):
    result_array.append(str(id_string))
    return result_array

def eval_collDay1(result_array, month, date_string):
    temp_date = str(date_string).split(" ")
    temp_date = temp_date[0].split("-")

    result_array.append(str(temp_date[2])) # Day
    result_array.append(str(month[int(temp_date[1])])) # Month
    result_array.append(str(temp_date[0])) # Year

    return result_array

def eval_collDay2(result_array, month, date_string):
    if date_string == None:
        result_array.append("-")
        result_array.append("-")
        result_array.append("-")

    elif len(str(date_string)) == 25:
        temp_date = str(date_string)[:10]
        temp_date = temp_date.split("-")

        result_array.append(str(temp_date[2])) # Day
        result_array.append(str(month[int(temp_date[1])])) # Month
        result_array.append(str(temp_date[0])) # Year

    else:
        result_array.append("manual fill")
        result_array.append("manual fill")
        result_array.append("manual fill")

    return result_array

def eval_collName(result_array, collector_array, code_string):
    # Convert the collector code to a matching Name from table
    match_flag = 0
    for c_row in collector_array:
        if c_row[0].lower() == str(code_string).lower():
            match_flag = 1
            result_array.append(str(c_row[1] + " " + c_row[2]))
            break

    if match_flag == 0:
        result_array.append("-")

    return result_array

def eval_collNum(result_array, num_string):
    result_array.append(str(num_string))
    return result_array

def eval_sampleNum(result_array, sample_num_flag, sample_num):
    if sample_num != None:
        if int(sample_num) > 1:
            # We want to create a row for each sample collected,
            # so we set the flag now to reference later
            sample_num_flag = 1
        result_array.append(1)
    else:
        # -1 represents an error or empty value
        result_array.append(-1)

    return result_array, sample_num_flag

def eval_state(result_array, num_string):
    result_array.append(str(num_string))
    return result_array

def eval_county(result_array, county_string):
    result_array.append(str(county_string))
    return result_array

def eval_city(result_array, city_string):
    if city_string == None:
        result_array.append("-")
    else:
        result_array.append(str(city_string))
    return result_array

def eval_location(result_array, loc_string):
    result_array.append(str(loc_string))
    return result_array

def eval_latLong(result_array, lat, long):
    # Lat
    if lat == None:
        result_array.append("-")
    else:
        result_array.append(str(round(lat,4)))
    # Long
    if long == None:
        result_array.append("-")
    else:
        result_array.append(str(round(long,4)))

    return result_array

def eval_colMethod(result_array, method_string):
    result_array.append(str(method_string))
    return result_array

def eval_assocPlant(result_array, plant_string):
    result_array.append(str(plant_string))
    return result_array

def eval_elevation(result_array, lat_string, long_string):
    # Use lat and long to look up elevation data
    elevation = 0;
    result_array.append(str(elevation))
    return result_array

def eval_positional_acc(result_array, accuracy_string):
    result_array.append(str(accuracy_string))
    return result_array

def eval_specimenID(result_array, id_string):
    result_array.append(str(id_string))
    return result_array

def eval_sampleID(result_array, id_string):
    result_array.append(str(id_string))
    return result_array

def eval_collectorID(result_array, id_string):
    result_array.append(str(id_string))
    return result_array

def eval_specimenURL(result_array, url_string):
    result_array.append(str(url_string))
    return result_array

def merge_tables(observation_array, collector_array, input_wb, input_ws, num_rows=None):
    # Load XLSX file
    wb = Workbook()
    ws = wb.active
    ws.title = input_ws

    # Initialize values
    month = ['i','ii','iii','iv','v','vi','vii','viii','ix','x','xi','xii']
    header_row = [  'iNaturalist ID',
                    'Collection Day 1',
                    'Month 1',
                    'Year 1',
                    'Collection Day 2',
                    'Month 2',
                    'Year 2',
                    'Collector Name',
                    'Collection No',
                    'Sample No',
                    'State',
                    'County',
                    'City',
                    'Location',
                    'Lat',
                    'Long',
                    'Collection method',
                    'Associated plant'
                ]
    ws.append(header_row)

    # Translate observation list rows to template format
    print("Translating iNaturalist list...")
    for row in observation_array:
        print(row)
        # Initialize values
        sample_num_flag = 0
        result_array = []

        # Begin constructing row to append to spreadsheet, saved in result_array

    # Col 1
    # Specimen ID
    # result_array = eval_specimenID(result_array, row[X])

        # Col 1
        # iNaturalist ID
        result_array = eval_iNaturalistID(result_array, row[0])

        # Col 2, 3, 4
        # Collection Day 1
        result_array = eval_collDay1(result_array, month, row[1])

        # Col 5, 6, 7
        # Collection Day 2
        result_array = eval_collDay2(result_array, month, row[17])

    # Col 1
    # Sample ID
    # result_array = eval_sampleID(result_array, row[X])

        # Col 8
        # Collector Name
        result_array = eval_collName(result_array, collector_array, row[2])

        # Col 9
        # Collection No
        result_array = eval_collNum(result_array, row[3])

        # Col 10
        # Sample No
        result_array, sample_num_flag = eval_sampleNum(result_array, sample_num_flag, row[4])

        # Col 11
        # State
        result_array = eval_state(result_array, row[6])

        # Col 12
        # County
        result_array = eval_county(result_array, row[7])

        # Col 13
        # City
        result_array = eval_city(result_array, row[15])

        # Col 14
        # Location
        result_array = eval_location(result_array, row[8])

        # Col 15, 16
        # Lat & Long
        result_array = eval_latLong(result_array, row[9], row[10])

    # Col X
    # Positional accuracy
    # result_array = eval_positional_acc(result_array, row[X])

    # Col X
    # Elevation
    # result_array = eval_positional_acc(result_array, row[X])

        # Col 17
        # Collection Method
        result_array = eval_colMethod(result_array, row[14])

        # Col 18
        # Associated Plant
        result_array = eval_assocPlant(result_array, row[12])

        # Append to xlsx file
        if sample_num_flag == 1:
            if row[4] != None:
                temp_range = row[4]
                for x in range(0, int(temp_range), 1):
                    print(result_array)
                    ws.append(result_array)
                    result_array[9] = int(result_array[9]) + 1
        else:
            print(result_array)
            ws.append(result_array)

    print("Appending results to file...")
    wb.save(filename = input_wb)
    print("Saved.")

def main():
    # Initiate / Default values
    input_folder = 'default'
    output_folder = 'default'
    min_col = 'A'
    min_row = '2'

    # Command line args
    print("cmd arguments: " , str(sys.argv))
    i = 0
    for arg in sys.argv:
        if arg == "--output":
            output_folder = sys.argv[i+1]
        elif arg == "--input":
            input_folder = sys.argv[i+1]
        i += 1

    # Default path settings
    observation_wb = 'data/' + input_folder + '/2018_iNaturalist.xlsx'
    observation_ws = ['observations-49204', 'Oregon Bee Atlas']

    # If output/input is set in cmd line, set the output folder to mirror that
    if output_folder != 'default':
        output_wb = 'results/' + output_folder + '/Oregon_Bee_Atlas.xlsx'
    elif input_folder != 'default':
        output_wb = 'results/' + input_folder + '/Oregon_Bee_Atlas.xlsx'
    else:
        output_wb = 'results/default/Oregon_Bee_Atlas.xlsx'
    output_ws = 'Results'

    # Create directories
    if not os.path.exists(os.path.dirname(output_wb)):
        try:
            os.makedirs(os.path.dirname(output_wb))
        except OSError as exc: # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise

    # Extract values from tables
    max_col = list(string.ascii_lowercase)[ int(count_cols(observation_wb,observation_ws[0])) - 1 ].upper()
    max_row = count_rows(observation_wb,observation_ws[0])
    observation_result = read_xlsx(observation_wb, observation_ws[0], min_col, min_row, max_col, max_row)

    max_col = list(string.ascii_lowercase)[ int(count_cols(observation_wb,observation_ws[1])) - 1 ].upper()
    max_row = count_rows(observation_wb,observation_ws[1])
    collector_result = read_xlsx(observation_wb, observation_ws[1], min_col, min_row, max_col, max_row)

    # Generate Output Sheet
    merge_tables(observation_result, collector_result, output_wb, output_ws)

if __name__ == '__main__':
    main()
