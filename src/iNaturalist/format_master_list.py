'''
Created on May 14th, 2018
Author: Miles McCall
Sources:
Description: Parse the "Pollinator Plant Master List" maintained by Signe and
            generate a formatted version for uploading to the website.
'''

# External Imports
import os
import sys
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import argparse

#Functions
def read_xlsx(wb_name, ws_name, min_col, min_row, max_col, max_row):
    # Initialize openpyxl vars
    wb = load_workbook(filename = str(wb_name))
    ws = wb[str(ws_name)]
    return_array = []

    # Iterate through xlsx file
    selection_str = min_col + min_row + ':' + max_col + max_row
    for row in ws.iter_rows(selection_str):
        temp_array = []
        for cell in row:
            temp_array.append(cell.value)
        return_array.append(temp_array)

    # Return array selection
    return return_array

'''def load_collector_name(input_array):
    wb = load_workbook(filename = str(input_wb))
    ws = wb[str(input_ws)]

    for row in master_array:
        result_array = []
        for col in columns:

    return collector_array
'''

def merge_tables(master_array, collector_arrayinput_wb, input_ws, num_rows=None):
    # Load XLSX file
    wb = load_workbook(filename = str(input_wb))
    ws = wb[str(input_ws)]

    header_row = [  'iNaturalist ID',
                    'Collection Day 1',
                    'Month 1',
                    'Year 1',
                    'Collector Name',
                    'Collection No',
                    'Sample No',
                    'State',
                    'County',
                    'Lat',
                    'Long',
                    'Collection method',
                    'Associated plant'
                ]

    columns = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N']

    month = ['i','ii','iii','iv','v','vi','vii','viii','ix','x','xi','xii']

    # Translate master list rows to template format
    print("Translating master list to XLSX format...")

    for row in master_array:
        sample_num_flag = 0
        result_array = []
        for col in columns:
            if(col == 'A'): # iNaturalist ID
                result_array.append(str(row[0]))

            elif(col == 'B'): # Collection Day 1
                date = str(row[2]).split("/")
                result_array.append(date[0]) # Day
                result_array.append(month(date[1]+1)) # Month
                result_array.append(str("20" + date[2])) # Year

            elif(col == 'C'): # Collector Name
                for c_row in collector_array:
                    if c_row[0].lower() == str(row[2]).lower():
                        result_array.append(str(c_row[1] + " " + c_row[2]))
                        break

            elif(col == 'D'): # Collection No
                result_array.append(str(row[3]))

            elif(col == 'E'): # Sample No
                if int(row[4]) > 1:
                    sample_num_flag = 1
                    result_array.append("1")
                else:
                    result_array.append(str(row[4]))

            elif(col == 'F'): # State
                result_array.append(str(row[6]))

            elif(col == 'G'): # County
                result_array.append(str(row[7]))

            elif(col == 'H'): # Lat
                result_array.append(str(round(row[9],4)))

            elif(col == 'I'): # Long
                result_array.append(str(round(row[10],4)))

            elif(col == 'J'): # Collection Method
                result_array.append(str(row[11]))

            elif(col == 'K'): # Associated Plant
                result_array.append(str(row[12]))

        # Append to xlsx file
        if sample_num_flag == 1:
            range = result_array[6]
            for x in range(0, range):
                print(result_array)
                ws.append(result_array)
                result_array[6] += 1
        else:
            print(result_array)
            ws.append(result_array)

    print("Appending results to input file...", end="")
    wb.save(input_wb)
    print("saved.")

def main():
    # Command Line Args
    parser = argparse.ArgumentParser(description='Web Template Formatter')
    parser.add_argument('--num_rows', dest='num_rows', type=int,
                       help='add this arg to set a max number of rows')
    args = parser.parse_args()
    print(str(args.num_rows))

    # Init master vars
    master_wb = 'data/2018_iNaturalist.xlsx'
    master_ws = ['observations-34528', 'Oregon Bee Atlas']

    # Extract values from tables
    min_col = 'A';  min_row = '2'
    max_col = 'P'; max_row = '252'
    master_res = read_xlsx(master_wb, master_ws[0], min_col, min_row, max_col, max_row)

    min_col = 'A';  min_row = '2'
    max_col = 'C'; max_row = '163'
    collector_res = read_xlsx(master_wb, master_ws[1], min_col, min_row, max_col, max_row)

    # Init template vars
    template_wb = 'data/Oregon_Bee_Atlas_Auto.xlsx'
    template_ws = 'Master'
    merge_tables(master_res, collector_res, template_wb, template_ws)

if __name__ == '__main__':
    main()
