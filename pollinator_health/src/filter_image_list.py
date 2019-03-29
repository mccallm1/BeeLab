'''
Created on May 14th, 2018
Author: Miles McCall
Sources:
Description: Pull all images from the copyright list missing sources
'''

# External Imports
import os
import sys
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
# Local Imports

#Functions
def read_xlsx(wb_name, ws_name, min_col, min_row, max_col, max_row):
    # Initialize openpyxl vars
    wb = load_workbook(filename = str(wb_name))
    ws = wb[str(ws_name)]
    return_array = []

    # Iterate through xlsx file
    selection_str = min_col + min_row + ':' + max_col + max_row
    for row in ws.iter_rows(selection_str):
        temp_array = []
        for cell in row:
            temp_array.append(cell.value)
        return_array.append(temp_array)

    # Return array selection
    return return_array

def translate_format(array, input_wb, input_ws):
    for row in array:
        if(row[8] == '' and row[9] == ''])


        result_array = []
        print(result_array)

    #ws.append(result_array)
    print("Appending results to input file...", end=" ")
    #wb.save(input_wb)
    print("saved.")

def main():
    # Init image vars
    image_wb = 'data/IPM_Copyright_Spreadsheet.xlsx'
    image_ws = 'IPM'
    min_col = 'A';  min_row = '2'
    max_col = 'K'; max_row = '204'
    # Extract values from master table
    image_res = read_xlsx(image_wb, image_ws, min_col, min_row, max_col, max_row)

    # Init template vars
    template_wb = 'data/plants_sample_data.xlsx'
    template_ws = 'Sample spreadsheet for the plan'

    translate_format(master_res, template_wb, template_ws)

if __name__ == '__main__':
    main()
